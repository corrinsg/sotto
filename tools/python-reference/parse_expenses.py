#!/usr/bin/env python3
"""Parse bank statement PDFs into a spreadsheet (CSV + Excel)."""

import argparse
import csv
import re
import sys
from pathlib import Path

import pdfplumber


def extract_transactions(pdf_path: str) -> list[dict]:
    """Extract transactions from a bank statement PDF."""
    transactions = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words:
                continue

            # Sort words by vertical position, then horizontal
            words.sort(key=lambda w: (round(w["top"], 1), w["x0"]))

            # Group words into lines by y-position
            lines = []
            current_line = []
            current_y = None
            for w in words:
                y = round(w["top"], 1)
                if current_y is None or abs(y - current_y) < 5:
                    current_line.append(w)
                    current_y = y if current_y is None else current_y
                else:
                    if current_line:
                        lines.append(current_line)
                    current_line = [w]
                    current_y = y
            if current_line:
                lines.append(current_line)

            # Find column boundaries from header row
            header_line = None
            for line in lines:
                text = " ".join(w["text"] for w in line)
                if "Paid out" in text or "paid out" in text.lower():
                    header_line = line
                    break

            if header_line is None:
                paid_out_right = 480
                paid_in_right = 620
                balance_right = 740
                details_x = 120
            else:
                # Use RIGHT edge (x1) of headers since amounts are right-aligned
                paid_words = [w for w in header_line if w["text"].lower() == "paid"]
                out_words = [w for w in header_line if w["text"].lower() == "out"]
                in_words = [w for w in header_line if w["text"].lower() == "in"]
                balance_words = [w for w in header_line if w["text"].lower() == "balance"]

                if out_words and in_words:
                    paid_out_right = out_words[0]["x1"]
                    paid_in_right = in_words[0]["x1"]
                elif len(paid_words) >= 2:
                    paid_out_right = paid_words[0]["x1"]
                    paid_in_right = paid_words[1]["x1"]
                elif len(paid_words) == 1:
                    paid_out_right = paid_words[0]["x1"]
                    paid_in_right = paid_out_right + 140
                else:
                    paid_out_right = 480
                    paid_in_right = 620

                balance_right = balance_words[0]["x1"] if balance_words else paid_in_right + 120
                details_x = 120

            # Column boundaries: midpoints between column right-edges
            boundary_out_in = (paid_out_right + paid_in_right) / 2
            boundary_in_bal = (paid_in_right + balance_right) / 2

            # Parse transaction lines (skip header and anything before it)
            header_y = header_line[-1]["top"] if header_line else 0
            data_lines = [l for l in lines if l[0]["top"] > header_y + 5]

            date_pattern = re.compile(r"^\d{1,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{2}$")
            amount_pattern = re.compile(r"^[\d,]+\.\d{2}$")

            current_txn = None

            for line in data_lines:
                line_text = " ".join(w["text"] for w in line)

                # Skip footer lines
                if any(kw in line_text for kw in ["BALANCE CARRIED", "continued", "Page"]):
                    if current_txn:
                        transactions.append(current_txn)
                        current_txn = None
                    continue

                # Check if this line starts with a date
                leftmost_x = line[0]["x0"]
                has_date = False
                date_str = ""

                # Check for date at the start (leftmost position, x < details_x)
                date_words_in_line = []
                other_words = []
                for w in line:
                    if w["x0"] < details_x - 20:
                        date_words_in_line.append(w)
                    else:
                        other_words.append(w)

                if date_words_in_line:
                    potential_date = " ".join(w["text"] for w in date_words_in_line)
                    if date_pattern.match(potential_date):
                        has_date = True
                        date_str = potential_date

                # Categorize remaining words by column position
                detail_words = []
                paid_out_val = ""
                paid_in_val = ""
                balance_val = ""

                words_to_check = other_words if date_words_in_line else line
                if has_date:
                    words_to_check = other_words

                for w in words_to_check:
                    text = w["text"]

                    if amount_pattern.match(text):
                        # Use right edge of the number to match right-aligned columns
                        right = w["x1"]
                        if right < boundary_out_in:
                            paid_out_val = text
                        elif right < boundary_in_bal:
                            paid_in_val = text
                        else:
                            balance_val = text
                    else:
                        detail_words.append(text)

                detail_text = " ".join(detail_words).strip()

                # Handle BALANCE BROUGHT FORWARD
                if "BALANCE BROUGHT FORWARD" in detail_text:
                    transactions.append({
                        "date": date_str,
                        "payment_type": "",
                        "details": "BALANCE BROUGHT FORWARD",
                        "paid_out": paid_out_val,
                        "paid_in": paid_in_val,
                        "balance": balance_val,
                    })
                    current_txn = None
                    continue

                # Detect payment type codes at the start of details
                payment_type = ""
                type_pattern = re.compile(r"^(VIS|BP|CR|DD|CHQ|SO|TFR|\)\)\)|FPI|FPO|BGC|ATM)\s*")
                m = type_pattern.match(detail_text)
                if m:
                    payment_type = m.group(1)
                    detail_text = detail_text[m.end():].strip()

                if has_date or payment_type:
                    # This starts a new transaction (or a new entry on the same date)
                    if current_txn:
                        transactions.append(current_txn)

                    current_txn = {
                        "date": date_str if has_date else "",
                        "payment_type": payment_type,
                        "details": detail_text,
                        "paid_out": paid_out_val,
                        "paid_in": paid_in_val,
                        "balance": balance_val,
                    }
                elif current_txn is not None:
                    # Continuation line — append details and pick up amounts
                    if detail_text:
                        current_txn["details"] += " " + detail_text
                    if paid_out_val and not current_txn["paid_out"]:
                        current_txn["paid_out"] = paid_out_val
                    elif paid_out_val:
                        current_txn["paid_out"] = paid_out_val
                    if paid_in_val and not current_txn["paid_in"]:
                        current_txn["paid_in"] = paid_in_val
                    elif paid_in_val:
                        current_txn["paid_in"] = paid_in_val
                    if balance_val and not current_txn["balance"]:
                        current_txn["balance"] = balance_val
                    elif balance_val:
                        current_txn["balance"] = balance_val

            if current_txn:
                transactions.append(current_txn)

    # Forward-fill dates
    last_date = ""
    for txn in transactions:
        if txn["date"]:
            last_date = txn["date"]
        else:
            txn["date"] = last_date

    return transactions


def write_csv(transactions: list[dict], output_path: str):
    """Write transactions to CSV."""
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["date", "payment_type", "details", "paid_out", "paid_in", "balance"],
        )
        writer.writeheader()
        writer.writerows(transactions)


def write_excel(transactions: list[dict], output_path: str):
    """Write transactions to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, numbers

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = ["Date", "Payment Type", "Details", "Paid Out (£)", "Paid In (£)", "Balance (£)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for txn in transactions:
        paid_out = float(txn["paid_out"].replace(",", "")) if txn["paid_out"] else None
        paid_in = float(txn["paid_in"].replace(",", "")) if txn["paid_in"] else None
        balance = float(txn["balance"].replace(",", "")) if txn["balance"] else None
        ws.append([txn["date"], txn["payment_type"], txn["details"], paid_out, paid_in, balance])

    # Format number columns
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "#,##0.00"

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Parse bank statement PDF into a spreadsheet")
    parser.add_argument("pdf", help="Path to the bank statement PDF")
    parser.add_argument("-o", "--output", help="Output file path (default: same name as input with .csv/.xlsx extension)")
    parser.add_argument("--format", choices=["csv", "xlsx", "both"], default="both", help="Output format (default: both)")
    args = parser.parse_args()

    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        print(f"Error: {pdf_path} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing {pdf_path}...")
    transactions = extract_transactions(str(pdf_path))
    print(f"Found {len(transactions)} transactions")

    if not transactions:
        print("No transactions found. The PDF format may not match the expected layout.", file=sys.stderr)
        sys.exit(1)

    stem = pdf_path.stem
    output_dir = pdf_path.parent

    if args.output:
        output_base = Path(args.output)
        stem = output_base.stem
        output_dir = output_base.parent or Path(".")

    if args.format in ("csv", "both"):
        csv_path = output_dir / f"{stem}.csv"
        write_csv(transactions, str(csv_path))
        print(f"CSV written to {csv_path}")

    if args.format in ("xlsx", "both"):
        xlsx_path = output_dir / f"{stem}.xlsx"
        write_excel(transactions, str(xlsx_path))
        print(f"Excel written to {xlsx_path}")


if __name__ == "__main__":
    main()
